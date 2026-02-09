import openpyxl 
import pandas
import os
from src.utils.load_yaml import ATTACHMENTS_SAVE_DIR,NLG_TEMPLATES_DIR,NLG_GENERATED_CENSUS_DIR,REFERRAL_FILE_STORE_DIR
from src.utils.support_functions import get_replaced_referral_id
from datetime import datetime


def nlg_map_census_data(id):
    census_filename = ""
    if id == 'default':
        for file_name in os.listdir(ATTACHMENTS_SAVE_DIR):
            if not file_name.startswith("Medical__") and file_name.endswith(".xlsx"):
                census_filename = file_name

        excel_data_df = pandas.read_excel(os.path.join(ATTACHMENTS_SAVE_DIR, census_filename), sheet_name='Sheet1')
        nationality_df = pandas.read_excel(os.path.join(ATTACHMENTS_SAVE_DIR, census_filename), sheet_name='Nationality_Updated')

    else:
        for file_name in os.listdir(os.path.join(REFERRAL_FILE_STORE_DIR, get_replaced_referral_id(id))):
            if not file_name.startswith("Medical__") and file_name.endswith(".xlsx"):
                census_filename = file_name

        excel_data_df = pandas.read_excel(os.path.join(REFERRAL_FILE_STORE_DIR, get_replaced_referral_id(id), census_filename), sheet_name='Sheet1')
        nationality_df = pandas.read_excel(os.path.join(REFERRAL_FILE_STORE_DIR, get_replaced_referral_id(id), census_filename), sheet_name='National_Updated')

    merged_df = pandas.merge(excel_data_df, nationality_df, left_on='Nationality', right_on='AL SAGR', how='left')
    merged_df['Salary Type']=merged_df['Salary Type'].apply(lambda x: 'Enhanced' if x == 'HSB' else 'LSB')
    merged_df['Relation']=merged_df['Relation'].apply(lambda x: 'Employee' if x == 'Principal' else x)
    merged_df['Visa Issued Emirates']=merged_df['Visa Issued Emirates'].apply(lambda x: 'DXB' if x == 'Dubai' else x)

    wb = openpyxl.load_workbook(os.path.join(NLG_TEMPLATES_DIR, "MemberUpload.xlsx"))
    ws = wb['loader']

    for index, row in merged_df.iterrows():
        ws.cell(row=index+2,column=1).value=row['Relation']
        ws.cell(row=index+2,column=2).value=row['Gender']
        # Convert the DOB to a datetime object if it's a string
        dob = row['DOB']
        if isinstance(dob, str):
            try:
                dob = datetime.strptime(dob.strip(), "%Y-%m-%d") 
            except ValueError:
                dob = None  
        if dob:
            ws.cell(row=index + 2, column=3).value = dob.strftime("%d-%b-%Y")
        else:
            ws.cell(row=index + 2, column=3).value = "Invalid DOB"
        ws.cell(row=index+2,column=4).value=row['Salary Type']
        ws.cell(row=index+2,column=5).value=row['Visa Issued Emirates']
        ws.cell(row=index+2,column=6).value='Cat ' + row['Category'] 
        ws.cell(row=index+2,column=7).value=row['Marital status']
        ws.cell(row=index+2,column=8).value=row['NLGIC Code']

    wb.save(os.path.join(NLG_GENERATED_CENSUS_DIR, "MemberUpload.xlsx"))