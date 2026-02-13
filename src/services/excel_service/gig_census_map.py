import openpyxl 
import pandas
import os
from src.utils.load_yaml import ATTACHMENTS_SAVE_DIR,GIG_TEMPLATES_DIR,REFERRAL_FILE_STORE_DIR, GIG_GENERATED_CENSUS_DIR
from src.utils.support_functions import get_replaced_referral_id
from datetime import datetime
import pandas as pd
from openpyxl.styles import NamedStyle
from src.services.excel_service.read_excel import read_excel
from src.utils.logger import logger


def parse_date_format(date_str):
    """
    Parse date string in various formats (dd/mm/yyyy, d/m/yyyy, dd-mm-yyyy, etc.)
    Returns a datetime object or None if parsing fails
    """
    if not isinstance(date_str, str):
        if isinstance(date_str, datetime):
            return date_str  # Already a datetime object
        return None
        
    date_str = date_str.strip()
    # Try multiple date formats
    formats = [
        "%d/%m/%Y",  # dd/mm/yyyy
        "%d-%m-%Y",  # dd-mm-yyyy
        "%d/%m/%y",  # dd/mm/yy
        "%d-%m-yy",  # dd-mm-yy
        "%m/%d/%Y",  # mm/dd/yyyy (in case it comes in US format)
        "%m-%d-%Y",  # mm-dd-yyyy
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
            
    logger.error(f"Could not parse date: {date_str}")
    return None


def gig_map_census_data(id, other_data=None):
    census_filename = ""
    if id == 'default':
        for file_name in os.listdir(ATTACHMENTS_SAVE_DIR):
            if not file_name.startswith("Medical__") and file_name.endswith(".xlsx"):
                census_filename = file_name
                break
        
        if not census_filename:
            raise FileNotFoundError("No suitable census file found in attachments directory")

        census_filepath = os.path.normpath(os.path.join(ATTACHMENTS_SAVE_DIR, census_filename))
        print(f"GIG Debug: Reading from {census_filepath}")
        
        if not os.path.exists(census_filepath):
            raise FileNotFoundError(f"Census file not found at: {census_filepath}")
            
        excel_data_df = pandas.read_excel(census_filepath, sheet_name='Sheet1')
        
        # Try to read nationality sheet, with fallback
        try:
            nationality_df = pandas.read_excel(census_filepath, sheet_name='Nationality_Updated')
        except ValueError:
            # Create a basic nationality mapping if sheet doesn't exist
            print("Warning: Nationality_Updated sheet not found, using basic mapping")
            unique_nationalities = excel_data_df['Nationality'].unique() if 'Nationality' in excel_data_df.columns else ['UAE']
            nationality_df = pandas.DataFrame({
                'AL SAGR': unique_nationalities,
                'GIG': unique_nationalities
            })

    else:
        for file_name in os.listdir(os.path.join(REFERRAL_FILE_STORE_DIR, get_replaced_referral_id(id))):
            if not file_name.startswith("Medical__") and file_name.endswith(".xlsx"):
                census_filename = file_name

        excel_data_df = pandas.read_excel(os.path.join(REFERRAL_FILE_STORE_DIR, get_replaced_referral_id(id), census_filename), sheet_name='Sheet1')
        nationality_df = pandas.read_excel(os.path.join(REFERRAL_FILE_STORE_DIR, get_replaced_referral_id(id), census_filename), sheet_name='Nationality_Updated')

    nationality_df.columns = nationality_df.columns.str.strip()
    merged_df = pandas.merge(excel_data_df, nationality_df, left_on='Nationality', right_on='AL SAGR', how='left')

    wb = openpyxl.load_workbook(os.path.join(GIG_TEMPLATES_DIR, "MemberUpload.xlsx"))
    ws = wb['Census']



    ##Effective Date - Read from database other_data instead of Excel file
    effective_date = None
    
    # Try to get effective date from other_data first - support multiple key formats
    if other_data and isinstance(other_data, dict):
        # Try multiple possible key formats for effective date
        possible_keys = ['Effective from', 'effective_from', 'effective_date', 'effectiveFrom']
        for key in possible_keys:
            effective_date = other_data.get(key)
            if effective_date:
                logger.info(f"Using effective date from database (key '{key}'): {effective_date}")
                break
        
        if not effective_date:
            logger.info("Effective date not found in database (tried keys: {}), trying to read from Excel file...".format(', '.join(possible_keys)))
    
    # Fallback to reading from Excel file if not found in other_data
    if not effective_date:
        if not other_data:  # Only show this message if no other_data was provided at all
            logger.info("Effective date not found in database, trying to read from Excel file...")
        df1, df2 = read_excel('GIG Insurance', 'default')
        
        # Check if the dataframes are empty
        if df1.empty:
            print("No data found sheet1")
            logger.error("No data found sheet1")
            raise Exception("No data found sheet1")
        
        # Extract the effective date from the DataFrame
        try:
            # Check if the required columns exist
            if 'KEY' in df1.columns and 'VALUE' in df1.columns:
                effective_from_rows = df1[df1['KEY'] == "Effective from"]
                if not effective_from_rows.empty and len(effective_from_rows['VALUE'].values) > 0:
                    effective_date = effective_from_rows['VALUE'].values[0].strip()
                    logger.info("Effective date read from Excel file: " + effective_date)
                else:
                    logger.warning("'Effective from' row not found in KEY column, using default date")
                    effective_date = "01/01/2024"  # Default date
            else:
                logger.warning("KEY or VALUE columns not found in df1, using default effective date")
                effective_date = "01/01/2024"  # Default date
        except Exception as e:
            logger.error(f"Error extracting effective date from Excel: {e}")
            effective_date = "01/01/2024"  # Default date
    
    # Ensure we have a valid effective date
    effective_date = effective_date or "01/01/2024"


    #Data mapping for census
    for index, row in merged_df.iterrows():
        
        #Name
        ws.cell(row=index+2, column=2).value = row['Beneficiary First Name']  

        #Member Type
        relation = row['Relation']
        if relation == 'Principal':
            relation = 'Employee'
        gender = row['Gender']
        marital_status = row['Marital status']

        if relation == 'Employee' :
            member_type=F'{relation} {gender} – {marital_status}'
        else :
            member_type=F'{relation} – {gender}'

        ws.cell(row=index+2, column=6).value = member_type

        #Relation
        if relation == 'Employee':
            relation = 'E' 
        elif relation == 'Spouse' and gender == 'Male':
            relation = 'H'
        elif relation == 'Spouse' and gender == 'Female':
            relation = 'W'
        elif relation == 'Child' and gender == 'Male':
            relation = 'S'
        elif relation == 'Child' and gender == 'Female':
            relation = 'D'
        ws.cell(row=index+2, column=3).value = relation

        #Gender
        if gender == 'Male':
            gender = 'M'
        elif gender == 'Female':
            gender = 'F'
        ws.cell(row=index+2, column=4).value = gender

        #Marital Status        
        if marital_status == 'Single':
            marital_status = 'S'
        elif marital_status == 'Married':
            marital_status = 'M'
        ws.cell(row=index+2, column=5).value = marital_status


        #Date of Birth
        dob = row['DOB']

        # Define a date format style (only needs to be done once)
        date_style = NamedStyle(name="date_style", number_format="DD-MM-YY")

        # Check if the style already exists to avoid duplicate errors
        if "date_style" not in ws.parent.named_styles:
            ws.parent.add_named_style(date_style)

        # Parse and set the DOB with flexible format handling
        dob_cell = ws.cell(row=index+2, column=7)
        if dob:
            parsed_dob = parse_date_format(dob)
            if parsed_dob:
                dob_cell.value = parsed_dob  # Set the datetime object
                dob_cell.number_format = "DD-MM-YY"  # Apply consistent date format
            else:
                dob_cell.value = "Invalid DOB"
                logger.warning(f"Invalid DOB format for row {index+2}: {dob}")
        else:
            dob_cell.value = "Invalid DOB"


        ##Nationality

        # Mappings with error handling
        if 'GIG INSURANCE' in nationality_df.columns and 'AL SAGR' in nationality_df.columns:
            nationality_mapping = dict(
                zip(nationality_df['AL SAGR'], nationality_df['GIG INSURANCE']))
        else:
            available_cols = nationality_df.columns.tolist()
            print(f"Warning: GIG INSURANCE column not found. Available columns: {available_cols}")
            if 'AL SAGR' in nationality_df.columns:
                nationality_mapping = dict(zip(nationality_df['AL SAGR'], nationality_df['AL SAGR']))
            else:
                nationality_mapping = {}
        
        # Fetch the mapped nationality; if not found, use the original value
        mapped_nationality = nationality_mapping.get(row['Nationality'], row['Nationality'])

        # Assign the mapped nationality to the Excel cell
        ws.cell(row=index+2, column=8).value = mapped_nationality
        # nationality = row['Nationality']
        # ws.cell(row=index+2, column=8).value = nationality
    

        ##Category
        cat=row['Category']
        if cat == 'A':
            cat = 'CAT 1'
        if cat == 'B':
            cat = 'CAT 2'
        if cat == 'C':
            cat = 'CAT 3'
        ws.cell(row=index+2, column=9).value = cat


    # Use the effective_date we determined above (from database or Excel fallback)
    logger.debug("Final effective date to use: " + str(effective_date))

    # Convert string to datetime object
    try:
        date_obj = datetime.strptime(effective_date, "%m/%d/%Y")  # Convert to datetime
    except ValueError:
        logger.error(f"Invalid date format: {effective_date}")
        date_obj = None  # Handle invalid date case

    # Define a date format style
    date_style = NamedStyle(name="date_style", number_format="DD/MM/YYYY")

    # Avoid duplicate named styles
    if "date_style" not in ws.parent.named_styles:
        ws.parent.add_named_style(date_style)

    # Select the cell where the date should be entered
    effective_date_cell = ws.cell(row=2, column=26)

    # Assign value and apply formatting
    if date_obj:
        effective_date_cell.value = date_obj  # Use datetime object so Excel recognizes it
        effective_date_cell.number_format = "DD/MM/YYYY"  # Ensure it's formatted as a date
        
        #**Force Excel to detect it as a date by reassigning**
        temp_value = effective_date_cell.value  # Store the value
        effective_date_cell.value = None  # Clear the cell
        effective_date_cell.value = temp_value  # Reassign the value

    else:
        effective_date_cell.value = "Invalid DOB"  # Handle errors properly

    logger.debug("Formatted effective date written to Excel.")
    
    # ws.cell(row=2, column=26).value = effective_date
    logger.debug(f"Effective Date: {effective_date}")

    wb.save(os.path.join(GIG_GENERATED_CENSUS_DIR, "gig_map.xlsx"))





