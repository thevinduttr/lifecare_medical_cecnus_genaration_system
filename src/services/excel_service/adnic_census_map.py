import openpyxl 
import pandas
import os
from src.utils.load_yaml import ATTACHMENTS_SAVE_DIR,ADNIC_TEMPLATES_DIR,ADNIC_GENERATED_CENSUS_DIR,REFERRAL_FILE_STORE_DIR
from src.utils.support_functions import get_replaced_referral_id
from datetime import datetime

def adnic_map_census_data(id):
    census_filename = ""
    if id == 'default':
        for file_name in os.listdir(ATTACHMENTS_SAVE_DIR):
            if not file_name.startswith("Medical__") and file_name.endswith(".xlsx"):
                census_filename = file_name
        excel_data_df = pandas.read_excel(os.path.join(ATTACHMENTS_SAVE_DIR, census_filename), sheet_name='Sheet1')
    else:
        for file_name in os.listdir(os.path.join(REFERRAL_FILE_STORE_DIR, get_replaced_referral_id(id))):
            if not file_name.startswith("Medical__") and file_name.endswith(".xlsx"):
                census_filename = file_name
        excel_data_df = pandas.read_excel(os.path.join(REFERRAL_FILE_STORE_DIR, get_replaced_referral_id(id), census_filename), sheet_name='Sheet1')

    excel_data_df['Salary Type']=excel_data_df['Salary Type'].apply(lambda x: 'Enhanced' if x == 'HSB' else 'LSB')
    excel_data_df['Relation'] = excel_data_df['Relation'].apply(lambda x: 'Employee' if x == 'Principal' else x)
    excel_data_df['Gender'] = excel_data_df['Gender'].apply(lambda x: 'M' if x == 'Male' else ('F' if x == 'Female' else x))
    excel_data_df['Visa Issued Emirates'] = excel_data_df['Visa Issued Emirates'].apply(lambda x: 'DXB' if x == 'Dubai' else x)

    wb = openpyxl.load_workbook(os.path.join(ADNIC_TEMPLATES_DIR, "MemberUpload.xlsx"))
    ws = wb['Sheet1']

    for index, row in excel_data_df.iterrows():
        ws.cell(row=index+2,column=1).value=row['Relation']
        ws.cell(row=index+2,column=2).value=row['Gender']
        # Convert the DOB to a datetime object if it's a string
        # dob = row['DOB']
        # if isinstance(dob, str):
        #     try:
        #         dob = datetime.strptime(dob.strip(), "%Y-%m-%d") 
        #     except ValueError:
        #         dob = None  
        # if dob:
        #     ws.cell(row=index + 2, column=3).value = dob.strftime("%d-%b-%y")
        # else:
        #     ws.cell(row=index + 2, column=3).value = "Invalid DOB"
            # Apply the DOB formatting
    
        dob = row['DOB']
        if pandas.notnull(dob):
            dob_converted = pandas.to_datetime(dob, dayfirst=True, errors='coerce')
            if pandas.notnull(dob_converted):
                formatted_dob = dob_converted.strftime("%d-%b-%y")  # Format changed to d-MMM-yy
 
            else:
                formatted_dob = "Invalid DOB"
        else:
            formatted_dob = "Invalid DOB"
 
        ws.cell(row=index+2,column=3).value = formatted_dob
        ws.cell(row=index+2,column=4).value=row['Salary Type']
        ws.cell(row=index+2,column=5).value=row['Visa Issued Emirates']
        ws.cell(row=index+2,column=6).value=row['Category'] 
        ws.cell(row=index+2,column=7).value=row['Marital status']

    wb.save(os.path.join(ADNIC_GENERATED_CENSUS_DIR, "MemberUpload.xlsx"))
