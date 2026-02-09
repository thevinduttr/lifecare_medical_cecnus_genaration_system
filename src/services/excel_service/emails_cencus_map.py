import pandas as pd
from datetime import datetime
import os
import openpyxl
import subprocess
import traceback

from src.utils.load_yaml import ATTACHMENTS_SAVE_DIR, EMAIL_CENCUS_TEMPLATE_DIR, EMAIL_GENERATED_CENSUS_DIR

def calculate_age(dob):
    today = datetime.today()
    return today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))

def email_map_census_data(id):
    try:
        # Find the census file
        census_filename = ""
        if id == 'default':
            for file_name in os.listdir(ATTACHMENTS_SAVE_DIR):
                if not file_name.startswith("Medical__") and file_name.endswith(".xlsx"):
                    census_filename = file_name
            
            census_filepath = os.path.join(ATTACHMENTS_SAVE_DIR, "CensusData-TEMPLATE_Common with Nationality.xlsx")
        else:
            raise ValueError("Only 'default' ID is supported in this implementation.")

        # Load the data from 'Sheet1' of the census file
        sheet1_data = pd.read_excel(census_filepath, sheet_name='Sheet1')

        # Debug: Print column names to verify
        print("Columns in the Excel file:", sheet1_data.columns.tolist())

        # Clean the 'Monthly salary' column (remove commas and convert to numeric)
        if 'Monthly salary' in sheet1_data.columns:
            sheet1_data['Monthly salary'] = sheet1_data['Monthly salary'].replace({',': ''}, regex=True).astype(float)
        else:
            raise KeyError(f"Column 'Monthly salary' not found in the Excel file. Available columns: {sheet1_data.columns.tolist()}")

        # Process the data for the new template
        new_structure = pd.DataFrame({
            'Sr.': [f"{i:03}" for i in range(1, len(sheet1_data) + 1)],  # Auto-incrementing Sr. column
            'Full Name': sheet1_data['Beneficiary First Name'],
            'Emirate of Visa Issuance': sheet1_data['Visa Issued Emirates'],
            'DOB': pd.to_datetime(sheet1_data['DOB']).dt.date,
            'Gender: Male/Female': sheet1_data['Gender'],
            'Marital Status: Married/Single': sheet1_data['Marital status'],
            'Nationality': sheet1_data['Nationality'],
            'Status: Employee / Spouse / Child': sheet1_data['Relation'],
            'Salary above 4K: Yes/No': ['Yes' if salary > 4000 else 'No' for salary in sheet1_data['Monthly salary']],
            'Category: A/B': sheet1_data['Category']
        })

        # Define the template file path
        template_filepath = os.path.join(EMAIL_CENCUS_TEMPLATE_DIR, "Lifecare_Email_Cencus_Template.xlsx")

        # Check if the template file exists
        if not os.path.exists(template_filepath):
            raise FileNotFoundError(f"Template file not found at: {template_filepath}. Please ensure the template exists.")

        # Load the template file using openpyxl
        workbook = openpyxl.load_workbook(template_filepath)
        sheet = workbook.active

        # Insert the processed data into the template starting from row 4
        for index, row in new_structure.iterrows():
            sheet.cell(row=index + 4, column=1, value=row['Sr.'])
            sheet.cell(row=index + 4, column=2, value=row['Full Name'])
            sheet.cell(row=index + 4, column=3, value=row['Emirate of Visa Issuance'])
            sheet.cell(row=index + 4, column=4, value=row['DOB'])
            sheet.cell(row=index + 4, column=5, value=row['Gender: Male/Female'])
            sheet.cell(row=index + 4, column=6, value=row['Marital Status: Married/Single'])
            sheet.cell(row=index + 4, column=7, value=row['Nationality'])
            sheet.cell(row=index + 4, column=8, value=row['Status: Employee / Spouse / Child'])
            sheet.cell(row=index + 4, column=9, value=row['Salary above 4K: Yes/No'])
            sheet.cell(row=index + 4, column=10, value=row['Category: A/B'])

        # Create the output directory if it doesn't exist
        os.makedirs(EMAIL_GENERATED_CENSUS_DIR, exist_ok=True)

        # Save the final file as Excel_Cencus.xlsx
        output_filepath = os.path.join(EMAIL_GENERATED_CENSUS_DIR, "Lifecare_Census Template.xlsx")
        workbook.save(output_filepath)

        # Close the workbook
        workbook.close()

        print(f"File saved at: {output_filepath}")

    except Exception as e:
        # Print the full error traceback
        print(f"An error occurred: {e}")
        traceback.print_exc()

# # Example usage
# email_map_census_data('default')