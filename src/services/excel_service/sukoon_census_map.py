import openpyxl
import pandas as pd
from datetime import datetime
from src.utils.load_yaml import SUKOON_GENERATED_CENSUS_DIR, SUKOON_TEMPLATES_DIR, ATTACHMENTS_SAVE_DIR, REFERRAL_FILE_STORE_DIR
from src.utils.support_functions import get_replaced_referral_id
import os


def sukoon_map_census_data(id):
    # File selection
    census_filename = ""
    if id == 'default':
        for file_name in os.listdir(ATTACHMENTS_SAVE_DIR):
            if not file_name.startswith("Medical__") and file_name.endswith(".xlsx"):
                census_filename = file_name
        excel_file_path = os.path.join(ATTACHMENTS_SAVE_DIR, census_filename)
    else:
        for file_name in os.listdir(os.path.join(REFERRAL_FILE_STORE_DIR, get_replaced_referral_id(id))):
            if not file_name.startswith("Medical__") and file_name.endswith(".xlsx"):
                census_filename = file_name
        excel_file_path = os.path.join(
            REFERRAL_FILE_STORE_DIR, get_replaced_referral_id(id), census_filename)

    if not census_filename:
        print("Error: Census file not found.")
        return

    print(f"Reading Excel file from {excel_file_path}")

    try:
        # Load input and nationality sheets
        input_df = pd.read_excel(excel_file_path, sheet_name="Sheet1")
        nationality_df = pd.read_excel(
            excel_file_path, sheet_name="Nationality_Updated")

        # Load output workbook and sheet
        output_wb = openpyxl.load_workbook(os.path.join(
            SUKOON_TEMPLATES_DIR, "MemberCensusDataTemplate.xlsx"))
        output_ws = output_wb["Medical Upload"]

    except (FileNotFoundError, ValueError, KeyError) as e:
        print(f"Error loading files or sheets: {e}")
        return

    # Mappings with error handling
    if 'SUKOON INSURANCE' in nationality_df.columns:
        nationality_mapping = dict(
            zip(nationality_df['AL SAGR'], nationality_df['SUKOON INSURANCE']))
    else:
        # Fallback: use available nationality column or direct mapping
        available_cols = nationality_df.columns.tolist()
        print(f"Warning: SUKOON INSURANCE column not found. Available columns: {available_cols}")
        if 'AL SAGR' in nationality_df.columns:
            nationality_mapping = dict(zip(nationality_df['AL SAGR'], nationality_df['AL SAGR']))
        else:
            nationality_mapping = {}
    
    column_mapping = {
        "Beneficiary First Name": "First Name ",
        "DOB": "Date of Birth (DD/MM/YYYY)",
        "Gender": "Gender",
        "Category": "Category",
        "Nationality": "Nationality",
        "Relation": "Relation",
        "Marital status": "Marital Status",
        "Visa Issued Emirates": "Region",
    }
    category_mapping = {'A': 1, 'B': 2, 'C': 3}
    
    # Only rename columns that exist in the dataframe
    existing_columns = {k: v for k, v in column_mapping.items() if k in input_df.columns}
    input_df = input_df.rename(columns=existing_columns)
    
    # Add missing Category column with default values if it doesn't exist
    if "Category" not in input_df.columns and "Category" not in existing_columns.values():
        input_df["Category"] = input_df.get('Status', 'A')  # Use Status or default to 'A'

    # Initialize employee number counter
    emp_no_counter = 1

    for index, row in input_df.iterrows():
        # Row setup
        row_num = index + 2
        output_ws.cell(row=row_num, column=1).value = index + 1  # SL No

        # Name handling
        first_name = row.get("First Name ", "")
        # Convert to string if it's not already a string
        if not isinstance(first_name, str):
            first_name = str(first_name) if pd.notnull(first_name) else ""
        name_parts = first_name.split() if first_name else []
        # First Name
        output_ws.cell(row=row_num, column=2).value = name_parts[0] if len(
            name_parts) >= 1 else "-"
        # Middle Name
        output_ws.cell(row=row_num, column=3).value = name_parts[1] if len(
            name_parts) > 2 else "-"
        # Last Name
        output_ws.cell(row=row_num, column=4).value = (name_parts[1] if len(name_parts) == 2 else " ".join(name_parts[2:]) if len(name_parts) > 2 else "-")

        # DOB formatting
        dob = row.get("Date of Birth (DD/MM/YYYY)")
        if pd.notnull(dob):
            try:
                formatted_dob = pd.to_datetime(dob, dayfirst=True).strftime(
                    "%d/%m/%Y") if not isinstance(dob, datetime) else dob.strftime("%d/%m/%Y")
            except Exception:
                formatted_dob = "Invalid DOB"
        else:
            formatted_dob = "Invalid DOB"
        output_ws.cell(row=row_num, column=6).value = formatted_dob

        # Gender, Marital Status, and Relation
        output_ws.cell(row=row_num, column=7).value = row.get("Gender", "")
        output_ws.cell(row=row_num, column=8).value = row.get(
            "Marital Status", "")

        relation = "Employee" if row.get(
            "Relation", "") == "Principal" else row.get("Relation", "")
        output_ws.cell(row=row_num, column=9).value = relation

        # Assign Employee Number
        output_ws.cell(
            row=row_num, column=5).value = emp_no_counter if relation == "Employee" else emp_no_counter - 1
        emp_no_counter += 1 if relation == "Employee" else 0

        # Category, Region, and LSB
        category_value = category_mapping.get(
            row.get("Category", ""), "Unknown")
        output_ws.cell(row=row_num, column=10).value = category_value
        output_ws.cell(row=row_num, column=11).value = row.get("Region", "")
        output_ws.cell(row=row_num, column=12).value = 1 if row.get(
            "LSB", "") == "HSB" else 2

        # Nationality mapping
        nationality = row.get("Nationality", "")
        output_ws.cell(row=row_num, column=13).value = nationality_mapping.get(
            nationality, "Unknown")

    # Save the output
    output_file_path = os.path.join(
        SUKOON_GENERATED_CENSUS_DIR, "MemberCensusData.xlsx")
    output_wb.save(output_file_path)
    print(f"Data successfully written to {output_file_path}")
