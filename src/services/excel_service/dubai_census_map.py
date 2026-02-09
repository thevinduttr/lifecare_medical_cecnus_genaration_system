import openpyxl
import pandas as pd
from datetime import datetime
from src.utils.load_yaml import ATTACHMENTS_SAVE_DIR, REFERRAL_FILE_STORE_DIR,DUBAIINSURANCE_GENERATED_CENSUS_DIR,DUBAIINSURANCE_TEMPLATES_DIR
from src.utils.support_functions import get_replaced_referral_id
import os

# File paths and sheet names
OUTPUT_FILE_PATH = os.path.join(DUBAIINSURANCE_GENERATED_CENSUS_DIR, "Dubaiinsurance_map.xlsx")
OUTPUT_SHEET_NAME = "loader"
INPUT_SHEET_NAME = "Sheet1"

def dubai_map_census_data(id):
    try:
        census_filename = ""
        if id == 'default':
            for file_name in os.listdir(ATTACHMENTS_SAVE_DIR):
                if not file_name.startswith("Medical__") and file_name.endswith(".xlsx"):
                    census_filename = file_name
            
            census_filepath = os.path.join(ATTACHMENTS_SAVE_DIR, census_filename)

        else:
            for file_name in os.listdir(os.path.join(REFERRAL_FILE_STORE_DIR, get_replaced_referral_id(id))):
                if not file_name.startswith("Medical__") and file_name.endswith(".xlsx"):
                    census_filename = file_name

            census_filepath = os.path.join(REFERRAL_FILE_STORE_DIR,get_replaced_referral_id(id), census_filename)

        # Load the Excel file
        census_filepath = os.path.join(ATTACHMENTS_SAVE_DIR, census_filename)

        # Load the data from 'Sheet1'
        input_df = pd.read_excel(census_filepath, sheet_name=INPUT_SHEET_NAME)

        print("Initial Input Data:")
        print(input_df.head())

    except FileNotFoundError:
        print(f"Error: The file does not exist.")
        return
    except ValueError:
        print(f"Error: The sheet {INPUT_SHEET_NAME} does not exist.")
        return

    # Replace 'Principal' with 'Employee' in the 'Relation' column
    input_df['Relation'] = input_df['Relation'].replace('Principal', 'Employee')
    print("After Replacing 'Principal' with 'Employee':")
    print(input_df.head())

    # Convert Salary Type to 'Yes' or 'No'
    input_df['Visa Issued Emirates'] = input_df['Visa Issued Emirates'].apply(lambda visa_location: 'DXB' if visa_location == 'Dubai' else ('No' if visa_location == 'AhuDubai' else visa_location))
    print("After Converting Salary Type:")
    print(input_df.head())

    # Rename columns as per template needs
    column_mapping = {
        "Relation": "Relation",
        "Gender": "Gender",
        "DOB": "DOB",
        "Category": "Category",
        "Marital status": "Marital Status",
        "Visa Issued Emirates": "Visa Location",
        "Salary Type": "Salary Type"
    }
    input_df.rename(columns=column_mapping, inplace=True)

    # Load the output workbook and sheet
    try:
        output_wb = openpyxl.load_workbook(os.path.join(DUBAIINSURANCE_TEMPLATES_DIR, "dubaiinsurance_template.xlsx"))
        output_ws = output_wb[OUTPUT_SHEET_NAME]
    except FileNotFoundError:
        print(f"Error: The file {OUTPUT_FILE_PATH} does not exist.")
        return
    except KeyError:
        print(f"Error: The sheet {OUTPUT_SHEET_NAME} does not exist in {OUTPUT_FILE_PATH}.")
        return

    # Apply the DOB formatting
    for index, row in input_df.iterrows():
        dob = row['DOB']
        if pd.notnull(dob):
            dob_converted = pd.to_datetime(dob, errors='coerce')
            if pd.notnull(dob_converted):
                formatted_dob = dob_converted.strftime("%d-%b-%y")  # Format changed to d-MMM-yy

            else:
                formatted_dob = "Invalid DOB"
        else:
            formatted_dob = "Invalid DOB"

        output_ws.cell(row=index + 2, column=3).value = formatted_dob
        output_ws.cell(row=index + 2, column=1).value = row["Relation"]
        output_ws.cell(row=index + 2, column=7).value = row["Marital Status"]
        output_ws.cell(row=index + 2, column=2).value = row["Gender"]
        # output_ws.cell(row=index + 2, column=7).value = row["Relation"]
        output_ws.cell(row=index + 2, column=4).value = "Enhanced"
        output_ws.cell(row=index + 2, column=5).value = row["Visa Location"]
        category_letter = row["Category"]
        category_mapping = {'A': "Category A", 'B': "Category B", 'C': "Category C"}
        category_value = category_mapping.get(category_letter, "Unknown")
        output_ws.cell(row=index + 2, column=6).value = category_value

    output_wb.save(OUTPUT_FILE_PATH)
    print(f"Data successfully written to {OUTPUT_FILE_PATH}")
