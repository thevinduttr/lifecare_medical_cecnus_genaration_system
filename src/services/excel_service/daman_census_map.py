import openpyxl
import pandas as pd
import os
import time
import win32com.client as win32  # Import the win32com.client module
from win32com.client import Dispatch
import logging
import shutil
import gc
import datetime
from src.utils.load_yaml import ATTACHMENTS_SAVE_DIR, DAMAN_TEMPLATES_DIR, REFERRAL_FILE_STORE_DIR, DAMAN_GENERATED_CENSUS_DIR
from src.utils.support_functions import get_replaced_referral_id

# Set up logging
logger = logging.getLogger(__name__)

def clear_excel_cache():
    """Clear the win32com gen_py cache to fix COM automation issues"""
    try:
        # Find the gen_py directory
        import win32com
        win32com_dir = os.path.dirname(win32com.__file__)
        gen_py_dir = os.path.join(win32com_dir, 'gen_py')
        
        # Check if it exists and delete it
        if os.path.exists(gen_py_dir):
            logger.info(f"Removing gen_py cache directory: {gen_py_dir}")
            shutil.rmtree(gen_py_dir)
            logger.info("Cache directory successfully removed.")
            return True
        return False
    except Exception as e:
        logger.warning(f"Failed to clear Excel cache: {e}")
        return False

def open_excel_file_safely(file_path):
    """Open an Excel file using a more reliable approach"""
    try:
        # Try direct dispatch first
        excel = Dispatch("Excel.Application")
        excel.Visible = False  # Keep hidden for better performance
        excel.DisplayAlerts = False  # Disable alerts
        workbook = excel.Workbooks.Open(file_path)
        return excel, workbook
    except Exception as e:
        logger.warning(f"Standard Excel opening failed: {e}")
        
        # Try clearing the cache and using DispatchEx
        clear_excel_cache()
        
        try:
            # Try with DispatchEx which creates a new Excel instance
            excel = Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            workbook = excel.Workbooks.Open(file_path)
            return excel, workbook
        except Exception as e2:
            logger.error(f"All methods to open Excel file failed: {e2}")
            raise e2

def daman_map_census_data(id, other_data=None):
    """
    Enhanced DAMAN census mapping with improved error handling and COM cleanup
    Now supports database-driven effective date like GIG mapper
    
    Args:
        id: Processing ID ('default' for main processing)
        other_data: Dictionary containing database configuration data
    """
    excel = None
    workbook = None
    
    try: 
        logger.info(f"Starting DAMAN census mapping for ID: {id}")
        request_data_df1 = None
        request_data_df2 = None
        
        if id == 'default':
            census_filename = ""
            request_filename = ""

            # Identify the required files in ATTACHMENTS_SAVE_DIR
            for file_name in os.listdir(ATTACHMENTS_SAVE_DIR):
                if file_name.startswith("Medical_") and file_name.endswith(".xlsx"):
                    request_filename = file_name  # File starting with 'Medical_'
                elif not file_name.startswith("Medical__") and file_name.endswith(".xlsx"):
                    census_filename = file_name  # File NOT starting with 'Medical__'

            logger.info(f"Census file: {census_filename}, Request file: {request_filename}")

            # Load the census and nationality data
            excel_data_df = pd.read_excel(os.path.join(ATTACHMENTS_SAVE_DIR, census_filename), sheet_name='Sheet1')
            
            # Try to load nationality sheet, create default if not found
            try:
                nationality_df = pd.read_excel(os.path.join(ATTACHMENTS_SAVE_DIR, census_filename), sheet_name='Nationality_Updated')
                logger.info("Nationality_Updated sheet loaded successfully")
            except ValueError:
                logger.warning("Nationality_Updated sheet not found, creating default mapping")
                unique_nationalities = excel_data_df['Nationality'].unique() if 'Nationality' in excel_data_df.columns else ['UAE']
                nationality_df = pd.DataFrame({
                    'AL SAGR': unique_nationalities,
                    'DAMAN': unique_nationalities  
                })

            # Try to load request data, create defaults if not found  
            try:
                if request_filename:
                    request_data_df1 = pd.read_excel(os.path.join(ATTACHMENTS_SAVE_DIR, request_filename), sheet_name='Sheet1')
                    request_data_df2 = pd.read_excel(os.path.join(ATTACHMENTS_SAVE_DIR, request_filename), sheet_name='Sheet2')
                    logger.info("Request data files loaded successfully")
                else:
                    raise FileNotFoundError("No request file found")
            except (ValueError, FileNotFoundError) as e:
                logger.warning(f"Request data files not found: {e}, using defaults")
                request_data_df1 = pd.DataFrame()
                request_data_df2 = pd.DataFrame({'Category': ['A'], 'Network': ['Default Network']})
            
            logger.info(f"Census data shape: {excel_data_df.shape}")
            print(excel_data_df.head())

        else:
            # Handle non-default case
            for file_name in os.listdir(os.path.join(REFERRAL_FILE_STORE_DIR, get_replaced_referral_id(id))):
                if not file_name.startswith("Medical__") and file_name.endswith(".xlsx"):
                    census_filename = file_name

            excel_data_df = pd.read_excel(os.path.join(REFERRAL_FILE_STORE_DIR, get_replaced_referral_id(id), census_filename), sheet_name='Sheet1')
            nationality_df = pd.read_excel(os.path.join(REFERRAL_FILE_STORE_DIR, get_replaced_referral_id(id), census_filename), sheet_name='National_Updated')

        # Merge with nationality data, handling missing columns gracefully
        try:
            if 'Nationality' in excel_data_df.columns and 'AL SAGR' in nationality_df.columns:
                merged_df = pd.merge(excel_data_df, nationality_df, left_on='Nationality', right_on='AL SAGR', how='left')
                logger.info("Data merged with nationality mapping successfully")
            else:
                merged_df = excel_data_df.copy()
                # Add DAMAN column if it doesn't exist
                if 'DAMAN' not in merged_df.columns:
                    merged_df['DAMAN'] = merged_df.get('Nationality', 'Unknown')
                logger.warning("Using original data without nationality merge")
        except Exception as e:
            logger.warning(f"Merge failed: {e}, using original data")
            merged_df = excel_data_df.copy()
            if 'DAMAN' not in merged_df.columns:
                merged_df['DAMAN'] = merged_df.get('Nationality', 'Unknown')

        # Load the template workbook using openpyxl
        template_path = os.path.join(DAMAN_TEMPLATES_DIR, "SME_Member_Details_Template.xlsx")
        logger.info(f"Loading template from: {template_path}")
        wb = openpyxl.load_workbook(template_path)
        ws = wb['Member_Details']
        
        # Extract Effective from date - Database first, then fall back to request data
        effective_from_date = None
        
        # First try to get from database (other_data) - support multiple key formats
        if other_data and isinstance(other_data, dict):
            # Try multiple possible key formats for effective date
            possible_keys = ['Effective from', 'effective_from', 'effective_date', 'effectiveFrom']
            for key in possible_keys:
                effective_from_date = other_data.get(key)
                if effective_from_date:
                    logger.info(f"Effective from date retrieved from database using key '{key}': {effective_from_date}")
                    ws['B16'] = effective_from_date
                    break
            
            if not effective_from_date:
                logger.info("No effective date found in database other_data (tried keys: {})".format(', '.join(possible_keys)))
        
        # Fall back to request data if not found in database
        if not effective_from_date and request_data_df1 is not None and 'KEY' in request_data_df1.columns and 'VALUE' in request_data_df1.columns:
            effective_from_row = request_data_df1[request_data_df1['KEY'] == 'Effective from']
            if not effective_from_row.empty:
                effective_from_date = effective_from_row.iloc[0]['VALUE']
                logger.info(f"Effective from date retrieved from request data: {effective_from_date}")
                ws['B16'] = effective_from_date
                
        if not effective_from_date:
            logger.warning("No effective from date found in database or request data")
                
        # Get unique categories from the 'Category' column with error handling
        if 'Category' in excel_data_df.columns:
            unique_categories = excel_data_df['Category'].unique()
        else:
            # If no Category column, use Status or create default categories
            logger.warning("No 'Category' column found, using Status column or default 'A'")
            if 'Status' in excel_data_df.columns:
                unique_categories = excel_data_df['Status'].unique()
            else:
                unique_categories = ['A']  # Default category

        logger.info(f"Processing {len(unique_categories)} unique categories: {unique_categories}")

        # Write the associated values (Salary Type, Visa Issued Emirates, Network) for each unique category to the Excel sheet
        for i, category in enumerate(unique_categories):
            # Extract associated values for the category
            category_column = 'Category' if 'Category' in excel_data_df.columns else ('Status' if 'Status' in excel_data_df.columns else None)
            if category_column:
                category_data = excel_data_df[excel_data_df[category_column] == category]
            else:
                category_data = excel_data_df.head(1)  # Use first row as default
            
            if len(category_data) > 0:
                salary_type = category_data['Salary Type'].iloc[0] if 'Salary Type' in category_data.columns else 'Default'
                visa_issued = category_data['Visa Issued Emirates'].iloc[0] if 'Visa Issued Emirates' in category_data.columns else 'UAE'
            else:
                salary_type = 'Default'
                visa_issued = 'UAE'

            # Extract the Network value for this category from request_data_df2
            try:
                if request_data_df2 is not None and 'Category' in request_data_df2.columns and 'Network' in request_data_df2.columns:
                    network_value = request_data_df2[request_data_df2['Category'] == category]['Network'].values
                    network = network_value[0] if len(network_value) > 0 else 'Default Network'
                else:
                    network = 'Default Network'  # Default if columns not found
            except Exception as e:
                logger.warning(f"Error extracting network for category {category}: {e}")
                network = 'Default Network'

            # Write the values into the Excel sheet (row 21+i, columns 2-5)
            ws.cell(row=21 + i, column=2).value = f"CAT {category}"  # Write Category (A, B, C, ...)
            ws.cell(row=21 + i, column=3).value = visa_issued  # Write Visa Issued Emirates for the category
            ws.cell(row=21 + i, column=4).value = network  # Write Network for the category
            ws.cell(row=21 + i, column=5).value = salary_type  # Write Salary Type for the category
            
            logger.debug(f"Category {category}: Visa={visa_issued}, Network={network}, Salary={salary_type}")

        # Write member data starting from row 52
        logger.info(f"Writing {len(merged_df)} member records")
        for index, row in merged_df.iterrows():
            ws.cell(row=index+52, column=1).value = row['Beneficiary First Name']
            ws.cell(row=index+52, column=2).value = row['DOB']
            ws.cell(row=index+52, column=3).value = row['Gender'] 
            ws.cell(row=index+52, column=4).value = row.get('DAMAN', row.get('Nationality', 'Unknown'))
            ws.cell(row=index+52, column=5).value = row['Relation']
            category_value = row.get('Category', row.get('Status', 'A'))
            ws.cell(row=index+52, column=6).value = f"CAT {category_value}"
            ws.cell(row=index+52, column=7).value = row['Visa Issued Emirates']

        # Save the workbook using openpyxl with file locking protection
        output_path = os.path.join(DAMAN_GENERATED_CENSUS_DIR, "SME_Member_Details_Template.xlsx")
        logger.info(f"Saving workbook to: {output_path}")
        
        # Check if file exists and is locked, remove it if possible
        if os.path.exists(output_path):
            try:
                # Try to remove the existing file to avoid permission issues
                os.remove(output_path)
                logger.info(f"Removed existing file: {output_path}")
            except PermissionError:
                logger.warning(f"Existing file is locked: {output_path}, attempting to force release...")
                # Force garbage collection to help release any file handles
                gc.collect()
                time.sleep(1)
                try:
                    os.remove(output_path)
                    logger.info(f"Successfully removed locked file after cleanup: {output_path}")
                except PermissionError as e:
                    logger.error(f"Cannot remove locked file: {e}")
                    # Try with a different filename
                    import datetime
                    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                    output_path = os.path.join(DAMAN_GENERATED_CENSUS_DIR, f"SME_Member_Details_Template_{timestamp}.xlsx")
                    logger.info(f"Using alternative filename: {output_path}")
        
        # Save with retry mechanism
        max_save_attempts = 3
        for attempt in range(max_save_attempts):
            try:
                wb.save(output_path)
                logger.info(f"Workbook saved successfully on attempt {attempt + 1}")
                break
            except PermissionError as e:
                logger.warning(f"Save attempt {attempt + 1} failed with permission error: {e}")
                if attempt < max_save_attempts - 1:
                    # Force cleanup and wait before retry
                    gc.collect()
                    time.sleep(2)
                else:
                    logger.error(f"All save attempts failed, raising exception")
                    raise e
            except Exception as e:
                logger.error(f"Unexpected error during save: {e}")
                raise e
                
        wb.close()  # Close openpyxl workbook

        # Open the Excel file using safer method for COM automation
        try:
            logger.info(f"Opening Excel file for COM processing: {output_path}")
            excel, workbook = open_excel_file_safely(output_path)

            # Brief wait to ensure file is ready
            time.sleep(2)

            # Save and close the workbook
            workbook.Save()
            logger.info("Excel file processed successfully")
            
        except Exception as excel_error:
            logger.error(f"Excel COM processing failed: {excel_error}")
            # Don't raise here as the file is already saved with openpyxl
            logger.info("Continuing without COM processing as openpyxl save was successful")
            
    except Exception as e:
        logger.error(f"Error in DAMAN census mapping: {e}")
        raise e
        
    finally:
        # Comprehensive COM cleanup
        try:
            if workbook:
                workbook.Close(SaveChanges=False)
                workbook = None
        except:
            pass
            
        try:
            if excel:
                excel.Quit()
                excel = None
        except:
            pass
        
        # Force garbage collection to release COM objects
        gc.collect()
        logger.info("DAMAN census mapping completed with cleanup")